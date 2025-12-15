#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script untuk mengekstrak konten dari file Word, Excel, dan PDF
ke format teks yang bisa dibaca oleh Cursor
"""

import os
import sys
from pathlib import Path

def extract_docx(file_path):
    """Ekstrak konten dari file Word (.docx)"""
    try:
        from docx import Document
        doc = Document(file_path)
        content = []
        for para in doc.paragraphs:
            if para.text.strip():
                content.append(para.text)
        return '\n'.join(content)
    except ImportError:
        return f"[ERROR: python-docx tidak terinstall. Install dengan: pip install python-docx]"
    except Exception as e:
        return f"[ERROR membaca file: {str(e)}]"

def extract_xlsx(file_path):
    """Ekstrak konten dari file Excel (.xlsx)"""
    try:
        import pandas as pd
        content = []
        xls = pd.ExcelFile(file_path)
        for sheet_name in xls.sheet_names:
            content.append(f"\n=== SHEET: {sheet_name} ===\n")
            df = pd.read_excel(xls, sheet_name=sheet_name)
            content.append(df.to_string())
        return '\n'.join(content)
    except ImportError:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            content = []
            for sheet_name in wb.sheetnames:
                content.append(f"\n=== SHEET: {sheet_name} ===\n")
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    if any(row_data):  # Skip empty rows
                        content.append('\t'.join(row_data))
            return '\n'.join(content)
        except ImportError:
            return f"[ERROR: pandas atau openpyxl tidak terinstall. Install dengan: pip install pandas openpyxl]"
    except Exception as e:
        return f"[ERROR membaca file: {str(e)}]"

def extract_pdf(file_path):
    """Ekstrak konten dari file PDF"""
    try:
        import PyPDF2
        content = []
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num, page in enumerate(pdf_reader.pages, 1):
                content.append(f"\n=== HALAMAN {page_num} ===\n")
                content.append(page.extract_text())
        return '\n'.join(content)
    except ImportError:
        try:
            import pdfplumber
            content = []
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    content.append(f"\n=== HALAMAN {page_num} ===\n")
                    content.append(page.extract_text())
            return '\n'.join(content)
        except ImportError:
            return f"[ERROR: PyPDF2 atau pdfplumber tidak terinstall. Install dengan: pip install PyPDF2 pdfplumber]"
    except Exception as e:
        return f"[ERROR membaca file: {str(e)}]"

def main():
    """Main function untuk mengekstrak semua file dokumen"""
    base_dir = Path('.')
    
    # Daftar file yang akan diekstrak
    files_to_extract = [
        '1. kriteria 1 infrastruktur.docx',
        '2. kritria 2 kemananan wahana.docx',
        '3. Kriteria 3 Keselamatan wahana.docx',
        '4. Kriteria 4 Kesehatan wahana.docx',
        '5. Kriteria 5 sistem Pengamanan wahana.docx',
        '6. Kriteria 6 informasi wahana.docx',
        'DATA OBWIS.docx',
        'SPESIFIKASI DATA OBWIS.docx',
        'Petunjuk pengisian check list wahana.docx',
        'RISK ASSESMENT OBJEK WISATA 2025.xlsx',
        'ASTINA ST TTG KE POLRES RA SMP OBJEK WISATA OK.pdf',
        'ST RA SMP OBJEK WISATA.pdf'
    ]
    
    print("Memulai ekstraksi dokumen...")
    print("=" * 60)
    
    for filename in files_to_extract:
        file_path = base_dir / filename
        if not file_path.exists():
            print(f"[SKIP] File tidak ditemukan: {filename}")
            continue
        
        print(f"\nMemproses: {filename}")
        
        # Tentukan output file berdasarkan ekstensi
        if filename.endswith('.docx'):
            content = extract_docx(file_path)
            output_file = base_dir / f"{Path(filename).stem}.txt"
        elif filename.endswith('.xlsx'):
            content = extract_xlsx(file_path)
            output_file = base_dir / f"{Path(filename).stem}.txt"
        elif filename.endswith('.pdf'):
            content = extract_pdf(file_path)
            output_file = base_dir / f"{Path(filename).stem}.txt"
        else:
            print(f"[SKIP] Format tidak didukung: {filename}")
            continue
        
        # Simpan ke file teks
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(f"FILE ASLI: {filename}\n")
                f.write("=" * 60 + "\n\n")
                f.write(content)
            print(f"[SUCCESS] Disimpan ke: {output_file}")
        except Exception as e:
            print(f"[ERROR] Gagal menyimpan: {str(e)}")
    
    print("\n" + "=" * 60)
    print("Ekstraksi selesai!")

if __name__ == '__main__':
    main()

